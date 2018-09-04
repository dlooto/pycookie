#coding=utf8

"""
used for file process
"""

import os, logging
from zipfile import BadZipFile

from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.worksheet import Worksheet

import settings
from openpyxl import Workbook, load_workbook

logger = logging.getLogger('django')


def is_file_exist(path):
    """ 判断文件或目录是否存在 """
    return os.path.exists(path)


def remove(path, fileName=None):
    """remove file from the filesystem"""
    if not fileName:
        fullpath = path
    else:
        fullpath = os.path.join(path, fileName)
        
    try:
        os.remove(fullpath)
        return True
    except OSError as e:
        logger.error("delete file %s error: %s" % (fullpath, e))
        return False


def save_file(file, base_dir, file_name):
    """保存文件
    @param file  传入的文件参数, request.FILES中获取的数据对象, file需要先经过rename处理, 以便获取到file.name
    @param file_name  保存的目标文件名
     """
    
    if not file: return ''
    try:
        dest = open('%s/%s/%s' % (settings.MEDIA_ROOT, base_dir, file_name), 'wb+')
        for chunk in file.chunks():
            dest.write(chunk)
        dest.close()
    except Exception as e:
        logger.exception(e)
        dest.close()

    return file_name


def upload_file(file, base_dir, stored_file_name):
    """

    :param file:
    :param base_dir:
    :param storge_file_name:
    :return: 由上传的文件名和
    """
    try:
        path = os.path.join(settings.MEDIA_ROOT, base_dir)
        if not os.path.exists(path):
            os.makedirs(path)
        file_path = '%s/%s' % (path, stored_file_name)
        destination = open(file_path, 'wb+')
        for chunk in file.chunks():
            destination.write(chunk)
        destination.close()
    except Exception as e:
        logger.info(e)
        return None
    return file.name, '%s%s' % (base_dir, stored_file_name)


# def batch_update_files(file_dict, base_dir):
#
#     if not file_dict:
#         return None
#     path = None
#     try:
#         path = os.path.join(settings.MEDIA_ROOT, base_dir)
#         if not os.path.exists(path):
#             os.makedirs(path)
#     except Exception as e:
#         logger.info(e)
#         return None
#     try:
#         for file, stored_file_name in file_dict.items():
#             file_path = '%s/%s' % (path, stored_file_name)
#             destination = open(file_path, 'wb+')
#             for chunk in file.chunks():
#                 destination.write(chunk)
#             destination.close()
#     except Exception as e:
#         logger.info(e)
#         return None
#
#     return file.name, '%s%s' % (base_dir, stored_file_name)






def get_file_extension(path):
    """
    获取文件扩展名
    :param path:文件名或文件路径
    :return: 返回文件扩展名
    """
    return os.path.splitext(path)[1]


class ExcelBasedOXL(object):
    """
    基于Openpyxl库的excel封装
    """

    @staticmethod
    def open_excel(path):
        """
        打开excel
        :param path: the path to open or a file-like object
        :type path: string or a file-like object open in binary mode c.f., :class:`zipfile.ZipFile`
        :return: 返回Workbook对象，即excel文件对象
        """
        if not path:
            return False, 'path参数不能为空'
        try:
            workbook = load_workbook(path, read_only=True, data_only=True)
            return True, workbook
        except (InvalidFileException, BadZipFile) as iex:
            logger.exception(iex)
            return False, '文件格式异常或文件损坏'
        except Exception as e:
            logger.exception(e)
            return False, '文件读取异常'

    @staticmethod
    def get_sheet(workbook, sheet_name):
        """
        获取Worksheet对象
        :return: 返回Worksheet对象
        """
        return workbook[sheet_name]

    @staticmethod
    def get_rows_len(sheet):
        """
        获取sheet的最大行数
        :param sheet: Worksheet对象
        :return: 返回行数
        """
        if not sheet:
            return None
        return sheet.max_row

    @staticmethod
    def get_cell_value(sheet, row, col):
        """
        获取单元格中内容
        :param sheet: Worksheet对象
        :param row: 行序列号，默认1是首行
        :param col: 列序列号，默认1是首列
        :return:
        """
        if not sheet:
            return None
        return sheet.cell(row, col).value

    @staticmethod
    def close(workbook):
        """
        关闭Workbook对象
        :return:
        """
        if workbook:
            workbook.close()

    @staticmethod
    def read_excel(wb, header_dict=None):
        """
        读取excel文件数据，带表头
        :param wb: Workbook对象
        :param header_dict: 表头字典，K:属性名；V:表头单元格数据
        :return:
            返回由读取结果（True/False）和读取的数据列表(excel_data)组成的Cuple
            excel_data是由多个sheet数据组成的List；
            sheet是由多个行数据组成的List；
            每行数据则是以header_dict的k作为键，对应单元格数据作为值组成的Dict

        """
        excel_data = []
        try:
            for sheet_name in wb.sheetnames:
                sheet_data = ExcelBasedOXL.read_sheet(wb[sheet_name], header_dict)
                if sheet_data:
                    excel_data.append(sheet_data)
            return True, excel_data
        except Exception as e:
            logger.exception(e)
            return False, "表头数据和指定的标准不一致或excel解析错误"

    @staticmethod
    def read_sheet(ws: Worksheet, header_dict=None):
        """
        读取单个sheet数据，带表头
        :param ws: Worksheet对象
        :param header_dict: 表头字典，K:键，一般为model属性；V:对应表头单元格数据
        :return: 返回一个List对象，该List对象由一组Dictionary对象构成
        """
        if not ws:
            return []

        sheet_data = []
        ws_rows_len = ws.max_row
        ws_column_len = ws.max_column

        # 读取不带表头的sheet
        if header_dict is None:
            for row in range(1, ws_rows_len + 1):
                row_data = []
                for col in range(1, ws_column_len + 1):
                    row_data.append(ws.cell(row=row, column=col).value)
                sheet_data.append(row_data)

        # 读取带表头的sheet
        ws_column_len = len(header_dict)

        # 读取首行数据
        header_data = []
        for col in range(1, ws_column_len+1):
            header_data.append(ws.cell(1, col).value)

        header_keys = []  # 表头顺序对应的关键字列表
        # 判断首行数据是否和指定的标准一致，如果一致，按顺序封装表头关键字；否则抛出异常
        for item in header_dict.items():
            for hdata in header_data:
                if item[1] == hdata:
                    header_keys.append(item[0])
                    break

        if len(header_keys) != len(header_dict):
            # 抛出“表单的表头数据和指定的标准不一致，请检查”异常
            raise Exception('ExcelHeaderNotMatched')

        # 读取业务数据
        for row in range(2, ws_rows_len+1):
            row_data = {}
            for col in range(1, ws_column_len+1):
                key = header_keys[col-1]
                value = ws.cell(row=row, column=col).value
                row_data[key] = value
            sheet_data.append(row_data)

        return sheet_data




